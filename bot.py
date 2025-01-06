from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from datetime import datetime, time, timedelta, timezone
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    JobQueue,
)
from openpyxl import Workbook, load_workbook
import os
import pytz
import logging

# Configure logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# Replace with your bot's API token
BOT_TOKEN = '7805113858:AAFWsiptVGKCWEvuXa3mB4jVHtlFBqbls5o'

# Path to the Excel file
EXCEL_FILE = "attendance_log.xlsx"

# Telegram group ID
GROUP_ID = -1002298567228  # Use negative for group IDs

# Set your local time zone (e.g., 'Europe/Paris' for UTC+1 or UTC+2)
LOCAL_TZ = pytz.timezone('Asia/Kolkata')  # Time zone for Chennai, Kolkata, Mumbai, New Delhi

# Function to initialize the Excel file if it doesn't exist
def initialize_excel_file():
    if not os.path.exists(EXCEL_FILE):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        # Add headers to the sheet
        sheet.append(["Date", "User Name", "User ID", "Attendance"])
        workbook.save(EXCEL_FILE)

# Function to display the attendance form in the group
async def send_form_to_group(context: ContextTypes.DEFAULT_TYPE):
    # Create the inline keyboard with "checkboxes"
    keyboard = [
        [
            InlineKeyboardButton("✅ Yes", callback_data="Yes"),
            InlineKeyboardButton("❌ No", callback_data="No"),
            InlineKeyboardButton("🌓 Half Day", callback_data="Half Day"),
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # Send the message with the inline keyboard to the group
    await context.bot.send_message(
    chat_id=GROUP_ID,
    text="*Are you working today?*",
    parse_mode="Markdown",
    reply_markup=reply_markup,
    )


# Function to schedule the daily form at a specific time
async def schedule_daily_form(context: ContextTypes.DEFAULT_TYPE):
    job_queue: JobQueue = context.job_queue

    # Define the target time for the job in local time
    target_time = time(hour=18, minute=22)  # This is your local time (18:03)

    # Convert local time to UTC
    local_now = datetime.now(LOCAL_TZ)
    target_datetime = LOCAL_TZ.localize(datetime.combine(local_now.date(), target_time))
    target_utc_time = target_datetime.astimezone(pytz.utc).time()

    # Schedule the job to run daily at the target UTC time
    job_queue.run_daily(
        send_form_to_group,
        target_utc_time,  # Pass the UTC time to the scheduler
    )
    logging.info(f"Scheduled daily attendance form at {target_datetime.strftime('%I:%M %p')} local time.")

# Function to handle button clicks
async def handle_button_click(update, context):
    query = update.callback_query
    await query.answer()  # Acknowledge the callback query

    user = query.from_user
    response = query.data  # Get the data from the button clicked
    current_time = datetime.now(LOCAL_TZ).strftime('%Y-%m-%d %I:%M:%S %p')  # 12-hour format with AM/PM

    # Append response to the Excel file
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    sheet.append([current_time, user.first_name, user.id, response])
    workbook.save(EXCEL_FILE)

    # Notify the user that their attendance has been recorded
    await query.edit_message_text(f"Response: {response}")

# Command to send test form manually
async def test_send_form(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await send_form_to_group(context)

# Main function to start the bot
def main():
    # Initialize the Excel file
    initialize_excel_file()

    # Configure the application with the bot's token
    application = ApplicationBuilder().token(BOT_TOKEN).build()

    # Add command handlers
    application.add_handler(CommandHandler("testform", test_send_form))

    # Add callback handler for button clicks
    application.add_handler(CallbackQueryHandler(handle_button_click))

    # Schedule the daily form
    application.job_queue.run_once(schedule_daily_form, 0)  # Initial call to schedule the job

    # Log a starting message
    logging.info("Bot is starting...")

    # Start the bot
    application.run_polling()

if __name__ == '__main__':
    main()
